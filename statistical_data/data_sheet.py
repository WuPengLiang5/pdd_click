import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import xlrd
import datetime

def initialize_table(filename):
    if not os.path.isfile(filename):
        # 文件不存在，创建新的Excel文件
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        # 在新的Excel文件中添加一些数据
        #worksheet.cell(row=1, column=1, value="Hello")
        #worksheet.cell(row=1, column=2, value="World")
        # 保存Excel文件
        workbook.save(filename)
    # 文件存在，打开已有的Excel文件
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active
    worksheet.cell(row=1,column=1,value="")
    for i in range(24):
        #单元格坐标
        cell_coordinates=xlref(1,i+2)
        # 获取单元格
        cell = worksheet[cell_coordinates]
        # 创建一个Alignment对象
        alignment = Alignment(horizontal='center', vertical='center')
        # 设置对齐方式
        cell.alignment = alignment
        worksheet.cell(row=1,column=i+2,value=i)
    workbook.save(filename)

    #对列进行切片
    #min_col=1,max_col=1，第一列的全部数据，在一个元组里
    #min_col=1,max_col=2，每一列分别放到一个元组里
    col_1 = list(workbook['Sheet1'].iter_cols(min_col=1,max_col=1,values_only=True))[0]
    is_insert = True
    current_time=getCurrentTime()
    for cell in col_1:
        if cell == current_time:
            print("excel“当前的时间的行”已存在：%s"%cell)
            is_insert=False
    if is_insert:
        total_rows=len(col_1)
        worksheet.cell(row=total_rows+1,column=1,value=current_time)
    workbook.save(filename)

    # 获取总行数
    #total_rows = worksheet.max_row
    #print('Excel表格的总行数为:', total_rows)
    
def current_row_col(filename):
    workbook = load_workbook(filename=filename)
    worksheet = workbook['Sheet1']
    
    row=None
    col=None
    
    #获取当前时间的行
    y_m_d=getCurrentTime()
    col_1=list(workbook['Sheet1'].iter_cols(min_col=1,max_col=1,values_only=True))[0]
    row=len(col_1)

    #获取当前小时的列
    h=getHour()
    row_1=list(workbook['Sheet1'].iter_rows(min_row=1,max_row=1,values_only=True))[0]
    for index in range(len(row_1)):
        if row_1[index] == h:
            col=index+1
            break

    return row,col
#获取某行后列的表格值，注：下标是从0开始的，最左上角为（0，0）
def ReadFromExcel(path,sheet,row,column):
    file = xlrd.open_workbook(path).sheet_by_name(sheet)    #打开指定路径下的Sheet1表
    value = file.cell_value(row-1,column-1)
    return value

#传参文件路径、行、列、填入值,注：下标从1开始，左上角坐标为（1，1）
def WriteToExcel(path,row,column,value):
    file = load_workbook(path)   #加载
    file_active = file.active   #激活
    
    cell_coordinates=xlref(row,column) #单元格坐标 
    cell = file_active[cell_coordinates] # 获取单元格   
    alignment = Alignment(horizontal='center', vertical='center') # 创建一个Alignment对象        
    cell.alignment = alignment# 设置对齐方式

    # 创建一个Border对象
    border = Border(left=Side(border_style='thick', color='00B050'),
                right=Side(border_style='thick', color='00B050'),
                top=Side(border_style='thick', color='00B050'),
                bottom=Side(border_style='thick', color='00B050'))
    # 设置单元格的边框
    cell.border = border
    
    file_active.cell(row,column,value)  #修改
    file.save(path) #保存
    print("写入成功！")
    
def getCurrentTime():
 
    # 获取当前时间
    now = datetime.datetime.now()
 
    # 将时间按指定格式转换为字符串
    formatted_time = now.strftime("%Y/%m/%d")
    return formatted_time

def getHour():
    now = datetime.datetime.now()
    return now.hour
    
def write_to_excel(filename = "data.xlsx",value="1,1"):
    
    initialize_table(filename)

    row,col=current_row_col(filename)
    
    WriteToExcel(filename,row,col,value)
    
def read_from_excel(filename = "data.xlsx",input_sheet = "Sheet1"):
    row,col=current_row_col(filename)
    return ReadFromExcel(filename,input_sheet,row,col)

def xlref(row, column, zero_indexed=False):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)

if __name__ == "__main__":
    #print(getCurrentTime())
    write_to_excel(value='1,2')
    value = read_from_excel()
    print(value)
    #print(xlref(1,2))
